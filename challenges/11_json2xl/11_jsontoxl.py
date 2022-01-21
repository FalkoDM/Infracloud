import json
import openpyxl
from openpyxl import workbook
from openpyxl.workbook.workbook import Workbook

# load the jsonfile into variable jsonfile and close it
with open("webex_groups.json") as file:
    jsonfile = json.load(file)
    file.close()

all_members = []
i = 0 
#print(type(jsonfile))
# for every group in the jsonfile select the member list and group name
for group in jsonfile["groups"]:
    member_list = jsonfile["groups"][i]["group"]["members"]
    group_name = jsonfile["groups"][i]["group"]["group_name"]
    i +=1
    for member in member_list: #for every member in the member list, add the key value group:group name
        member.update({'group': group_name})
        all_members.append(member) # append it to an empty list

# print(all_members)
# print(type(all_members))

wb = Workbook() # create workbook
sheet = wb.active

# name columns
sheet.cell(1,1, "person_name")
sheet.cell(1,2, "email")
sheet.cell(1,3, "group_name")
row = 1
# populate rows 
for record in all_members:
    #print(record["person_name"])
    row +=1
    sheet.cell(row, 1, record["person_name"])
    sheet.cell(row, 2, record["email"])
    sheet.cell(row, 3, record["group"])
wb.save("test_json.xlsx") # save it to an excel file
    
